"""
reporter.py

Computes pipeline metrics and writes the two deliverables:
  - outputs/standardized_products.xlsx : one row per record after
    normalization and standardization, with a cluster id linking
    detected duplicates.
  - outputs/exception_report.xlsx : records that need a human to look
    at them (ambiguous fuzzy matches, missing fields, unmapped brands,
    category mismatches), plus a metrics summary sheet.

Excel formatting is kept plain on purpose: Calibri 11, thin header
border, no fills or icon sets. This is meant to read like an internal
ops report, not a dashboard mockup.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd


HEADER_FONT = Font(name="Calibri", size=11, bold=True)
BODY_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(bottom=Side(style="thin", color="B7B7B7"))


def _write_dataframe_sheet(ws, df):
    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center")

    for row in df.itertuples(index=False):
        ws.append(list(row))

    for i, col in enumerate(df.columns, start=1):
        max_len = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str).tolist()[:2000]])
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 10), 45)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT

    ws.freeze_panes = "A2"


def compute_metrics(total_records, standardized_df, duplicate_clusters,
                     auto_merged_pairs, review_pairs, exception_df,
                     unmapped_brand_count, category_mismatch_count,
                     missing_field_count):
    records_in_clusters = sum(len(members) for members in duplicate_clusters.values())
    unique_after_dedup = total_records - (records_in_clusters - len(duplicate_clusters))

    metrics = {
        "total_records_processed": total_records,
        "duplicate_clusters_found": len(duplicate_clusters),
        "records_involved_in_duplicates": records_in_clusters,
        "estimated_unique_products_after_dedup": unique_after_dedup,
        "duplicate_rate_pct": round(100 * (records_in_clusters - len(duplicate_clusters)) / total_records, 2) if total_records else 0,
        "auto_merged_pairs": len(auto_merged_pairs),
        "pairs_flagged_for_review": len(review_pairs),
        "records_in_exception_report": len(exception_df),
        "exception_rate_pct": round(100 * len(exception_df) / total_records, 2) if total_records else 0,
        "unmapped_brand_count": unmapped_brand_count,
        "category_mismatch_count": category_mismatch_count,
        "missing_field_count": missing_field_count,
    }
    return metrics


def write_standardized_output(path, standardized_df):
    wb = Workbook()
    ws = wb.active
    ws.title = "standardized_products"
    _write_dataframe_sheet(ws, standardized_df)
    wb.save(path)


def write_exception_report(path, exception_df, review_pairs_df, metrics):
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "summary"
    ws_summary.append(["Metric", "Value"])
    for cell in ws_summary[1]:
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    for key, value in metrics.items():
        ws_summary.append([key.replace("_", " "), value])
    for row in ws_summary.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
    ws_summary.column_dimensions["A"].width = 38
    ws_summary.column_dimensions["B"].width = 18

    ws_exceptions = wb.create_sheet("records_for_review")
    _write_dataframe_sheet(ws_exceptions, exception_df)

    ws_pairs = wb.create_sheet("ambiguous_match_pairs")
    _write_dataframe_sheet(ws_pairs, review_pairs_df)

    wb.save(path)
